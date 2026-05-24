import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import json
import io
from datetime import datetime

# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="AutoForm™ v2.0 | 자동 처방설계",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stAppViewContainer"] > .main { background: #EDF2F7; }
[data-testid="stSidebar"] { background: #1E2A3A; }
[data-testid="stSidebar"] * { color: #E2E8F0 !important; }
[data-testid="stSidebar"] .stButton button { background: #4A5568; color: white; border: none; }

.main-header {
    background: #1E2A3A;
    color: white;
    padding: 22px 32px;
    border-radius: 8px;
    margin-bottom: 22px;
    border-bottom: 3px solid #2B6CB0;
}
.section-label {
    background: #1E2A3A;
    color: white;
    padding: 10px 18px;
    border-radius: 6px 6px 0 0;
    font-weight: 700;
    font-size: 13px;
    letter-spacing: 0.4px;
    margin-bottom: 0;
}
.section-body {
    background: white;
    border: 1px solid #CBD5E0;
    border-top: none;
    border-radius: 0 0 6px 6px;
    padding: 20px 22px;
    margin-bottom: 18px;
}
.stitle {
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #1E2A3A;
    border-bottom: 2px solid #1E2A3A;
    padding-bottom: 6px;
    margin: 22px 0 14px 0;
}
.rat-box {
    background: #EDF2F7;
    border-left: 3px solid #1E2A3A;
    padding: 11px 15px;
    margin-bottom: 10px;
    border-radius: 0 4px 4px 0;
    font-size: 13px;
    line-height: 1.7;
}
.pstep-wrap { display: flex; flex-wrap: wrap; align-items: center; gap: 4px; margin: 8px 0 16px; }
.pstep { background: #1E2A3A; color: white; padding: 6px 13px; border-radius: 4px; font-size: 11px; font-weight: 700; }
.parrow { color: #4A5568; font-size: 16px; font-weight: bold; }
.sumcard {
    background: white;
    border: 1px solid #CBD5E0;
    border-left: 5px solid #1E2A3A;
    border-radius: 6px;
    padding: 14px 16px;
}
.sumcard .lbl { font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: .5px; color: #4A5568; }
.sumcard .val { font-size: 18px; font-weight: 700; color: #1E2A3A; margin-top: 3px; }
.sumcard .unit { font-size: 11px; color: #718096; }
div[data-testid="stMetric"] { background: white; border: 1px solid #CBD5E0; border-left: 5px solid #1E2A3A; border-radius: 6px; padding: 12px 16px; }
</style>
""", unsafe_allow_html=True)

# ─── SESSION STATE ────────────────────────────────────────────────────────────
if 'history' not in st.session_state:
    st.session_state.history = []
if 'result_data' not in st.session_state:
    st.session_state.result_data = None

# ─── FORMULATION ENGINES ──────────────────────────────────────────────────────

def engine_dc(p, api_pct):
    comp, warn, rat, suit = [], [], [], []
    w = p['weight']
    bcs = p['bcs']
    hygro = p['hygro']

    lub = 0.5
    glid = 0.2

    # 붕해제
    if bcs == 'II':
        dn, dg, dp = 'Crospovidone (PVPP)', 'Kollidon CL-M / Polyplasdone XL-10', 5
        dr = 'BCS II 저용해도 보완: PVPP 탄성회복+모세관 이중 기전 5%. CCS 대비 소수성 약물 용출 개선에 유리 (Schlack & Bhatt, 2020).'
    elif bcs == 'IV':
        dn, dg, dp = 'Crospovidone (PVPP)', 'Kollidon CL-M / Polyplasdone XL-10', 5
        dr = 'BCS IV: 빠른 붕해가 용출 개선의 첫 단계. PVPP 5% 선택 + 가용화 전략 병행 필수.'
        warn.append('⚠️ BCS IV: 직타 단독으로 용출 달성 어려울 수 있음. SLS 0.5% 추가 또는 고체분산체 기술 검토 강력 권고.')
    else:
        dn, dg, dp = 'Croscarmellose Sodium (CCS)', 'Ac-Di-Sol SD-711 / Primellose', 3
        dr = f'BCS {bcs} 표준 붕해제. CCS 팽윤(80%)+모세관(20%) 복합 기전, 권장 2–5% (USP <1151>). Lactose 배합성 우수.'

    # 윤활제
    if bcs in ('II', 'IV') or hygro:
        ln, lg = 'Sodium Stearyl Fumarate (NaSF)', 'PRUV / Lubripharm SSF'
        lr = f'BCS {bcs}/흡습성: NaSF 선택으로 Mg stearate 소수성 film에 의한 용출 지연 방지. 친수성, 혼합 시간 비민감 (Bolhuis & Armstrong, 2006).'
    else:
        ln, lg = 'Magnesium Stearate', 'HyQual / Ligamed MF-2-V (Veg.)'
        lr = 'Mg stearate 0.5% (0.25–1.0% 권장). Veg. grade로 BSE/TSE 위험 회피. 혼합 최종 2–3분 제한.'

    has_surf = (bcs == 'IV')
    dil = 100 - api_pct - lub - glid - dp - (0.5 if has_surf else 0)
    if dil < 5:
        warn.append(f'⚠️ 희석제 공간 부족 ({dil:.1f}%). 정제 중량 증가 필요.')
    dil = max(dil, 0.1)

    # 희석제
    if api_pct < 30:
        d1n, d1g, r1, d2n, d2g, r2 = 'Microcrystalline Cellulose','Avicel PH-101 / Vivapur 101',0.60,'Lactose Monohydrate','Pharmatose 200M / Tablettose 80',0.40
        rat.append({'t':'희석제 (저함량)', 'b':f'API {api_pct:.1f}% — MCC PH-101(60%): 압축성·내재 붕해능 우수. Lactose(40%): 유동성·외관 개선. MCC:Lac=6:4 DC 표준 (HoPE 8th Ed.).'})
    elif api_pct <= 60:
        d1n, d1g, r1, d2n, d2g, r2 = 'Microcrystalline Cellulose','Avicel PH-102 / Vivapur 102',0.40,'Lactose Monohydrate','Pharmatose 200M / Tablettose 80',0.60
        rat.append({'t':'희석제 (중간 함량)', 'b':f'API {api_pct:.1f}% — PH-102(입경 100μm) 유동성 우수 + Lactose. 압축성-유동성 균형 최적화.'})
    else:
        d1n, d1g, r1 = 'Dibasic Calcium Phosphate Dihydrate','Di-Cafos A150 / Emcompress', 0.70 if hygro else 0.75
        d2n, d2g, r2 = 'Microcrystalline Cellulose','Avicel PH-102', 0.30 if hygro else 0.25
        rat.append({'t':'희석제 (고함량)', 'b':f'API {api_pct:.1f}% — DCPD(고압축성, Aw<0.20) + MCC 취성 보완. 고함량 DC 표준.'})
        warn.append(f"⚠️ API {api_pct:.1f}%: DC 전 Carr's Index ≤25%, Hausner Ratio ≤1.35 확인 필요. 미충족 시 WG 전환 권고.")

    if p.get('ultra_low'): warn.append('⚠️ 극저함량(<1mg): 기하학적 희석법, Blend uniformity study 필수.')
    if p.get('poor_flow'): warn.append('⚠️ 유동성 불량: Aerosil 증량 또는 WG/DG 전환 검토.')

    comp += [
        {'step':'Intragranular','name':p['api_name'],'fn':'주성분','grade':'기준규격','mg':p['api_dose'],'pct':api_pct},
        {'step':'Intragranular','name':d1n,'fn':'희석제','grade':d1g,'mg':w*dil*r1/100,'pct':dil*r1},
        {'step':'Intragranular','name':d2n,'fn':'희석제','grade':d2g,'mg':w*dil*r2/100,'pct':dil*r2},
        {'step':'Intragranular','name':dn,'fn':'붕해제','grade':dg,'mg':w*dp/100,'pct':dp},
    ]
    if has_surf:
        comp.append({'step':'Intragranular','name':'Sodium Lauryl Sulfate (SLS)','fn':'가용화제','grade':'Kolliphor SLS Fine / EMPICOL LQ33','mg':w*0.5/100,'pct':0.5})
    comp += [
        {'step':'Extragranular','name':'Colloidal Silicon Dioxide','fn':'활택보조제','grade':'Aerosil 200 / Cab-O-Sil M-5P','mg':w*glid/100,'pct':glid},
        {'step':'Extragranular','name':ln,'fn':'윤활제','grade':lg,'mg':w*lub/100,'pct':lub},
    ]
    rat += [{'t':'붕해제 선택','b':dr},{'t':'윤활제 선택','b':lr}]
    if has_surf: rat.append({'t':'가용화제 추가 (BCS IV)','b':'SLS 0.5%: 위장관 약물 습윤화·미셀 형성. CMC≈6–8mM 이하 안전. 1% 초과 시 점막 자극 주의 (ICH Q3C).'})

    suit = [
        {'l':'DC 적합성','s':'pass' if api_pct<=70 else 'warn','d':f'API {api_pct:.1f}%: ' + ('표준 범위' if api_pct<=70 else '>70% — 유동성·압축성 확인 필수')},
        {'l':'용출 위험도','s':'pass' if bcs in ('I','III') else 'warn','d':f'BCS Class {bcs}'},
        {'l':'부형제 안전성','s':'pass','d':'ICH Q3C / FDA GRAS 목록 내'},
        {'l':'흡습성 대응','s':'warn' if hygro else 'pass','d':'RH≤45% 제조 환경 권고' if hygro else '이상 없음'},
        {'l':'광감수성 대응','s':'warn' if p.get('light') else 'pass','d':'차광 포장 필요' if p.get('light') else '이상 없음'},
        {'l':'정제 중량','s':'pass' if 100<=w<=800 else 'warn','d':f'{w}mg — ' + ('표준 범위' if 100<=w<=800 else '범위 확인 필요')},
    ]
    steps = ['원료 칭량','체과 (40 mesh)','Diluent+API 혼합','붕해제 혼합 (10-15min)','Glidant 혼합 (3min)','윤활제 혼합 (2-3min)','타정','(코팅 — 선택)']
    return {'comp':comp,'warn':warn,'rat':rat,'suit':suit,'steps':steps,'compat':build_compat(p,'DC')}


def engine_wg(p, api_pct):
    comp, warn, rat, suit = [], [], [], []
    w = p['weight']
    bcs = p['bcs']
    hygro = p['hygro']
    heat = p.get('heat', False)

    bp, ig_d, eg_d, lub, glid, eg_mcc = 3, 2, 2, 0.5, 0.2, 10
    ig_dil = 100 - api_pct - bp - ig_d - eg_d - lub - glid - eg_mcc
    if ig_dil < 5:
        warn.append(f'⚠️ IG 희석제 부족 ({ig_dil:.1f}%). 정제 중량 증가 또는 EG-MCC 감소 검토.')
    ig_dil = max(ig_dil, 0.1)

    if heat:
        bn, bg, br = 'Hydroxypropylcellulose (HPC-LF)','Klucel LF / Nisso HPC-L','열감수성: HPC-LF 저온(35–45°C) 건조 가능. 에탄올 용액 사용 가능, 열분해 위험 최소화.'
    elif hygro:
        bn, bg, br = 'HPMC (E5)','Methocel E5 / Pharmacoat 603','흡습성 API: HPMC E5는 PVP 대비 낮은 흡습성. 과립 표면 보호.'
    else:
        bn, bg, br = 'Povidone (PVP K-30)','Kollidon 30 / Plasdone K-29/32','WG 표준 결합제 PVP K-30 3%. 5–10% 수용액. 광범위 호환성·재현성 우수 (Rowe et al., HoPE 8th Ed.).'

    if hygro:
        igdn, igdg = 'Dibasic Calcium Phosphate Dihydrate','Di-Cafos A150 / Emcompress'
    else:
        igdn, igdg = 'Lactose Monohydrate','Pharmatose 200M / Tablettose 80'

    if bcs in ('II','IV'):
        din, dig = 'Crospovidone (PVPP)','Kollidon CL-M / Polyplasdone XL-10'
        dir_ = f'BCS {bcs}: PVPP IG 2%+EG 2% 분할. 이중 붕해 전략 (Parikh, 2016).'
    else:
        din, dig = 'Croscarmellose Sodium (CCS)','Ac-Di-Sol SD-711 / Primellose'
        dir_ = 'CCS 총 4% IG/EG 분할(50:50). 과립 내부+과립 간 이중 붕해 달성.'

    comp += [
        {'step':'Intragranular','name':p['api_name'],'fn':'주성분','grade':'기준규격','mg':p['api_dose'],'pct':api_pct},
        {'step':'Intragranular','name':igdn,'fn':'희석제 (IG)','grade':igdg,'mg':w*ig_dil/100,'pct':ig_dil},
        {'step':'Intragranular','name':bn,'fn':'결합제','grade':bg,'mg':w*bp/100,'pct':bp},
        {'step':'Intragranular','name':din+' (IG)','fn':'붕해제 (IG)','grade':dig,'mg':w*ig_d/100,'pct':ig_d},
        {'step':'Extragranular','name':'Microcrystalline Cellulose','fn':'희석제 (EG)','grade':'Avicel PH-102 / Vivapur 102','mg':w*eg_mcc/100,'pct':eg_mcc},
        {'step':'Extragranular','name':din+' (EG)','fn':'붕해제 (EG)','grade':dig,'mg':w*eg_d/100,'pct':eg_d},
        {'step':'Extragranular','name':'Colloidal Silicon Dioxide','fn':'활택보조제','grade':'Aerosil 200','mg':w*glid/100,'pct':glid},
        {'step':'Extragranular','name':'Magnesium Stearate','fn':'윤활제','grade':'HyQual / Ligamed MF-2-V (Veg.)','mg':w*lub/100,'pct':lub},
    ]
    rat += [
        {'t':'WG 선택 근거','b':'결합제 용액으로 과립화 → 유동성·압축성 개선. 고API 함량/유동성 불량 처방에 최적. 붕해제 IG/EG 분할(50:50)로 이중 붕해 전략 (Parikh, 2016).'},
        {'t':'결합제 선택','b':br},
        {'t':'붕해제 분할 전략','b':dir_},
        {'t':'EG MCC 추가','b':'EG MCC PH-102 10%: 과립화 후 압축성 저하 보완. 과립-MCC 경계면 소성변형 완충 역할. 5–15% 범위 조정 가능.'},
    ]
    suit = [
        {'l':'WG 적합성','s':'pass','d':'유동성·압축성 불량 API 최적 공정'},
        {'l':'열감수성 대응','s':'warn' if heat else 'pass','d':'건조 ≤45°C 권고' if heat else '표준 건조(50–60°C) 가능'},
        {'l':'수분 안정성','s':'warn' if hygro else 'pass','d':'에탄올 대체 또는 RH≤40% 권고' if hygro else '이상 없음'},
        {'l':'용출 위험도','s':'pass' if bcs in ('I','III') else 'warn','d':f'BCS Class {bcs}'},
        {'l':'부형제 안전성','s':'pass','d':'GRAS / USP-NF 등재'},
        {'l':'공정 복잡도','s':'warn','d':'DC 대비 단계 증가 (과립화+건조+정립). Scale-up 재현성 검증 필요.'},
    ]
    steps = ['원료 칭량','IG 혼합 (Pre-blend)','결합제 용액 제조','습식과립화 (High-Shear / Fluid Bed)','유동층 건조 (40–60°C)','정립 (18–24 mesh)','EG 혼합','윤활제 혼합 (2–3min)','타정']
    return {'comp':comp,'warn':warn,'rat':rat,'suit':suit,'steps':steps,'compat':build_compat(p,'WG')}


def engine_dg(p, api_pct):
    comp, warn, rat, suit = [], [], [], []
    w = p['weight']
    bcs = p['bcs']

    ig_d, eg_d, lub_ig, lub_eg, glid = 2, 3, 0.5, 0.5, 0.3
    dil = 100 - api_pct - ig_d - eg_d - lub_ig - lub_eg - glid
    if dil < 5: warn.append(f'⚠️ 희석제 부족 ({dil:.1f}%). 정제 중량 증가 검토.')
    dil = max(dil, 0.1)

    if bcs in ('II','IV'):
        din, dig, dir_ = 'Crospovidone (PVPP)','Kollidon CL-M','DG BCS II/IV: PVPP 건식 조건에서도 팽윤 특성 유지. IG/EG 분할.'
    else:
        din, dig, dir_ = 'Croscarmellose Sodium','Ac-Di-Sol SD-711','CCS IG 2%+EG 3%. 롤러압밀 후 붕해성 저하 보완 위해 EG 비율 증가.'

    comp += [
        {'step':'Intragranular','name':p['api_name'],'fn':'주성분','grade':'기준규격','mg':p['api_dose'],'pct':api_pct},
        {'step':'Intragranular','name':'Microcrystalline Cellulose','fn':'희석제 (IG)','grade':'Avicel PH-200 / Prosolv SMCC 90','mg':w*dil*0.55/100,'pct':dil*0.55},
        {'step':'Intragranular','name':din+' (IG)','fn':'붕해제 (IG)','grade':dig,'mg':w*ig_d/100,'pct':ig_d},
        {'step':'Intragranular','name':'Magnesium Stearate (IG)','fn':'윤활제 (IG)','grade':'HyQual','mg':w*lub_ig/100,'pct':lub_ig},
        {'step':'Extragranular','name':'Lactose Monohydrate','fn':'희석제 (EG)','grade':'Pharmatose 200M','mg':w*dil*0.45/100,'pct':dil*0.45},
        {'step':'Extragranular','name':din+' (EG)','fn':'붕해제 (EG)','grade':dig,'mg':w*eg_d/100,'pct':eg_d},
        {'step':'Extragranular','name':'Colloidal Silicon Dioxide','fn':'활택보조제','grade':'Aerosil 200','mg':w*glid/100,'pct':glid},
        {'step':'Extragranular','name':'Magnesium Stearate (EG)','fn':'윤활제 (EG)','grade':'HyQual','mg':w*lub_eg/100,'pct':lub_eg},
    ]
    rat += [
        {'t':'DG 선택 근거','b':'롤러압밀: 결합제 없이 물리적 압밀. 수분·열 민감 API 최적. MCC PH-200은 DG 후 잔류 압축성(residual compressibility) 우수.'},
        {'t':'IG 윤활제 역할','b':'IG Mg stearate: 롤러 점착(sticking) 방지. EG Mg stearate: 최종 타정 윤활. 합계 1.0%. EG 혼합 시간 ≤2분 제한 필수.'},
        {'t':'붕해제 분할','b':dir_},
    ]
    suit = [
        {'l':'수분·열 민감성 대응','s':'pass','d':'무용매 과립화 가능'},
        {'l':'롤러압밀 장비','s':'warn','d':'전용 Roller Compactor 필요'},
        {'l':'과립 밀도 관리','s':'warn','d':'과밀압밀 시 붕해 지연. Ribbon density 모니터링'},
        {'l':'용출 위험도','s':'pass' if bcs in ('I','III') else 'warn','d':f'BCS Class {bcs}'},
        {'l':'부형제 안전성','s':'pass','d':'GRAS / USP-NF'},
        {'l':'공정 재현성','s':'warn','d':'롤러 파라미터 최적화 및 Scale-up 검증 필요'},
    ]
    steps = ['원료 칭량','IG 혼합 (5–10min)','IG 윤활 (2min)','롤러압밀 (Roller Compaction)','정립 (Milling / 체과)','EG 혼합 (5–10min)','Glidant+윤활제 혼합 (2min)','타정']
    return {'comp':comp,'warn':warn,'rat':rat,'suit':suit,'steps':steps,'compat':build_compat(p,'DG')}


def engine_cap(p, api_pct):
    comp, warn, rat, suit = [], [], [], []
    w = p['weight']
    bcs = p['bcs']
    hygro = p['hygro']

    di, lub, glid = 4, 0.5, 0.5
    dil = 100 - api_pct - di - lub - glid
    if dil < 10: warn.append(f'⚠️ 충전 공간 부족 ({dil:.1f}%). 캡슐 크기 증가 검토.')
    dil = max(dil, 0.1)

    comp += [
        {'step':'충전물','name':p['api_name'],'fn':'주성분','grade':'기준규격','mg':p['api_dose'],'pct':api_pct},
        {'step':'충전물','name':'Microcrystalline Cellulose','fn':'희석제','grade':'Avicel PH-101 / PH-102','mg':w*dil*0.65/100,'pct':dil*0.65},
        {'step':'충전물','name':'Lactose Monohydrate','fn':'희석제','grade':'Pharmatose 200M','mg':w*dil*0.35/100,'pct':dil*0.35},
        {'step':'충전물','name':'Croscarmellose Sodium','fn':'붕해제','grade':'Ac-Di-Sol','mg':w*di/100,'pct':di},
        {'step':'충전물','name':'Colloidal Silicon Dioxide','fn':'유동화제','grade':'Aerosil 200','mg':w*glid/100,'pct':glid},
        {'step':'충전물','name':'Magnesium Stearate','fn':'윤활제','grade':'HyQual','mg':w*lub/100,'pct':lub},
        {'step':'캡슐 쉘','name':'HPMC 경질캡슐' if hygro else '젤라틴 경질캡슐','fn':'용기(Shell)','grade':'Vcaps Plus / Capsugel HPMC' if hygro else 'Capsugel DB Caps / ACG Gelatin','mg':76,'pct':0},
    ]
    rat += [
        {'t':'캡슐 선택 근거','b':'압축성 불량·극소 함량 API에 적합. 타정 단계 없어 API 물리적 스트레스 최소화.'},
        {'t':'캡슐 쉘 선택','b':'HPMC: 수분 흡수 ≤6% (vs 젤라틴 13–16%). 흡습성 API 보호.' if hygro else '젤라틴 경질캡슐: 표준 등급. 약물 방출 예측 용이, 비용 효율적.'},
        {'t':'희석제 선택','b':'MCC/Lactose(65:35): 분말 유동성·충전 균일성 개선. MCC는 충전 중 마찰 감소·점착 방지.'},
    ]
    suit = [
        {'l':'압축 단계 없음','s':'pass','d':'API 물리적 스트레스 최소화'},
        {'l':'충전 유동성','s':'warn','d':"Carr's Index ≤20% 권장. 불량 시 과립화 후 충전"},
        {'l':'흡습성 대응','s':'pass','d':'HPMC 캡슐 선택' if hygro else '표준 캡슐 적용 가능'},
        {'l':'함량 균일성','s':'pass' if api_pct<70 else 'warn','d':'모니터링 강화 필요' if api_pct>=70 else '표준 혼합으로 달성 가능'},
        {'l':'캡슐 크기 적합성','s':'warn','d':'충전 중량에 따라 최종 결정 필요 (Size 5~000)'},
        {'l':'부형제 안전성','s':'pass','d':'GRAS / USP-NF 등재'},
    ]
    steps = ['원료 칭량','체과 (40 mesh)','혼합 (Bin blender, 10–15min)','윤활제 혼합 (2min)','캡슐 충전 (Automatic Capsule Filler)','무게 검사','(Band Sealing)']
    return {'comp':comp,'warn':warn,'rat':rat,'suit':suit,'steps':steps,'compat':build_compat(p,'CAP')}


def build_compat(p, method):
    bcs = p['bcs']
    hygro = p['hygro']
    heat = p.get('heat', False)
    rows = [
        {'성분 조합':'API — Lactose','위험도':'중간' if hygro else '낮음','고려사항':'Maillard 반응 가능(아민계). 호환성 시험 권고.' if hygro else '호환성 우수. 아민계 약물 추가 확인 권고.'},
        {'성분 조합':'API — Mg Stearate','위험도':'낮음','고려사항':'소수성 막 형성: 혼합 2–3분 제한. 과혼합 시 dissolution 저하.'},
        {'성분 조합':'API — Aerosil (SiO₂)','위험도':'낮음','고려사항':'대부분 약물 호환성 우수. 고흡착성 약물은 AUC 영향 확인.'},
    ]
    if method == 'WG':
        rows.insert(2, {'성분 조합':'API — 결합제 용액','위험도':'중간' if heat else '낮음','고려사항':'열감수성: LOD 1–3%, 건조 온도 최소화.' if heat else '건조 중 API 안정성 확인 권고.'})
    if bcs == 'IV':
        rows.append({'성분 조합':'API — SLS (계면활성제)','위험도':'중간','고려사항':'이온성 약물 이온쌍 형성 가능. pH별 용출 패턴 모니터링 필요.'})
    return rows


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────
def export_excel(params, result, api_pct):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "처방설계"

    charcoal = "1E2A3A"
    slate = "4A5568"
    bg = "EDF2F7"
    alt = "F7FAFC"
    bdr_c = "CBD5E0"
    pass_c = "276749"
    warn_c = "B7791F"
    fail_c = "9B2335"

    thin = lambda: Border(
        left=Side(style='thin', color=bdr_c), right=Side(style='thin', color=bdr_c),
        top=Side(style='thin', color=bdr_c), bottom=Side(style='thin', color=bdr_c)
    )

    type_map = {'DC':'직접압축 (DC)','WG':'습식과립 (WG)','DG':'건식과립 (DG)','CAP':'캡슐 충전'}

    def hdr(cell, val, fg, font_color='FFFFFF', bold=True, size=10, align='center'):
        cell.value = val
        cell.fill = PatternFill(fill_type='solid', fgColor=fg)
        cell.font = Font(name='Arial', bold=bold, size=size, color=font_color)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = thin()

    def dat(cell, val, fg='FFFFFF', color='2D3748', bold=False, align='left'):
        cell.value = val
        cell.fill = PatternFill(fill_type='solid', fgColor=fg)
        cell.font = Font(name='Arial', bold=bold, size=10, color=color)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = thin()

    # Title
    ws.merge_cells('A1:F1')
    hdr(ws['A1'], f"AutoForm™ v2.0  —  완제의약품 자동 처방설계 결과", charcoal, size=13)
    ws.row_dimensions[1].height = 28

    # Info row
    ws.merge_cells('A2:F2')
    info = f"API: {params['api_name']}  |  함량: {params['api_dose']}mg / {params['weight']}mg  |  API 비율: {api_pct:.1f}%  |  BCS: Class {params['bcs']}  |  제형: {type_map[params['form_type']]}  |  생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    hdr(ws['A2'], info, slate, size=9, align='left')
    ws.row_dimensions[2].height = 18

    # Table header
    cols = ['공정단계', '성분명', '기능', '규격 / 참고 제품명', 'mg/정', '비율(%)']
    for ci, col in enumerate(cols, 1):
        hdr(ws.cell(3, ci), col, charcoal)
    ws.row_dimensions[3].height = 22

    # Data
    comp = result['comp']
    for i, c in enumerate(comp):
        row = i + 4
        is_shell = c['step'] == '캡슐 쉘'
        row_bg = alt if i % 2 == 0 else 'FFFFFF'
        dat(ws.cell(row,1), c['step'], row_bg, align='center')
        dat(ws.cell(row,2), c['name'], row_bg)
        dat(ws.cell(row,3), c['fn'], row_bg, align='center')
        dat(ws.cell(row,4), c['grade'], row_bg, color='4A5568')
        dat(ws.cell(row,5), f"{c['mg']:.2f}" if not is_shell else f"{c['mg']:.0f} (별도)", row_bg, align='right', bold=True)
        dat(ws.cell(row,6), f"{c['pct']:.2f}%" if not is_shell else '—', row_bg, align='right')

    # Total
    tr = len(comp) + 4
    total_mg = sum(c['mg'] for c in comp if c['step'] != '캡슐 쉘')
    total_pct = sum(c['pct'] for c in comp if c['step'] != '캡슐 쉘')
    ws.merge_cells(f'A{tr}:D{tr}')
    hdr(ws.cell(tr,1), '합계 (Total)', bg, font_color='1E2A3A', align='right')
    hdr(ws.cell(tr,5), f'{total_mg:.1f}', bg, font_color='1E2A3A', align='right')
    hdr(ws.cell(tr,6), f'{total_pct:.1f}%', bg, font_color='1E2A3A', align='right')
    ws.row_dimensions[tr].height = 20

    # Suitability section
    sr = tr + 2
    ws.merge_cells(f'A{sr}:F{sr}')
    hdr(ws.cell(sr,1), '공정 적합성 평가 (Process Suitability Assessment)', charcoal, align='left', size=11)
    for ci, h in enumerate(['평가 항목','판정','비고 (상세)'], 1):
        hdr(ws.cell(sr+1, ci), h, slate)
    if ci < 3: ws.merge_cells(f'C{sr+1}:F{sr+1}')
    ws.merge_cells(f'C{sr+1}:F{sr+1}')

    for i, s in enumerate(result['suit']):
        row = sr + 2 + i
        c_map = {'pass': pass_c, 'warn': warn_c, 'fail': fail_c}
        l_map = {'pass': '✓ PASS', 'warn': '⚠ CAUTION', 'fail': '✗ FAIL'}
        fg_map = {'pass': 'F0FFF4', 'warn': 'FFFBEB', 'fail': 'FFF5F5'}
        dat(ws.cell(row,1), s['l'], fg_map[s['s']])
        hdr(ws.cell(row,2), l_map[s['s']], fg_map[s['s']], font_color=c_map[s['s']])
        ws.merge_cells(f'C{row}:F{row}')
        dat(ws.cell(row,3), s['d'], alt, color='4A5568')

    # Compatibility
    cr = sr + len(result['suit']) + 3
    ws.merge_cells(f'A{cr}:F{cr}')
    hdr(ws.cell(cr,1), '주요 호환성 고려사항 (Compatibility Considerations)', charcoal, align='left', size=11)
    for ci, h in enumerate(['성분 조합','위험도','고려사항 및 권고'], 1):
        hdr(ws.cell(cr+1, ci), h, slate)
    ws.merge_cells(f'C{cr+1}:F{cr+1}')
    for i, c in enumerate(result['compat']):
        row = cr + 2 + i
        r_c = {'낮음': pass_c, '중간': warn_c, '높음': fail_c}
        dat(ws.cell(row,1), c['성분 조합'], alt if i%2==0 else 'FFFFFF', bold=True)
        hdr(ws.cell(row,2), c['위험도'], alt if i%2==0 else 'FFFFFF', font_color=r_c.get(c['위험도'], '2D3748'))
        ws.merge_cells(f'C{row}:F{row}')
        dat(ws.cell(row,3), c['고려사항'], alt if i%2==0 else 'FFFFFF', color='4A5568')

    # Disclaimer
    dr_row = cr + len(result['compat']) + 3
    ws.merge_cells(f'A{dr_row}:F{dr_row}')
    disc = ws.cell(dr_row, 1)
    disc.value = "※ 본 처방은 ICH Q8(R2), USP <1151>, HoPE 8th Ed. 기반 초기 설계 제안(Draft)입니다. 실제 개발 시 호환성 시험·DoE 최적화·Scale-up·안정성 시험이 필수입니다."
    disc.fill = PatternFill(fill_type='solid', fgColor=bg)
    disc.font = Font(name='Arial', size=9, color='4A5568', italic=True)
    disc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    disc.border = thin()

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── HEADER ──────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <div style="font-size:28px;font-weight:900;letter-spacing:3px">
    Auto<span style="color:#63B3ED">Form</span>™
    <span style="font-size:13px;font-weight:400;color:#90CDF4;letter-spacing:1px"> v2.0</span>
  </div>
  <div style="font-size:11px;color:#90CDF4;margin-top:5px">완제의약품 자동 처방설계 시스템 · Pharmaceutical Formulation Design Engine</div>
  <div style="font-size:10px;color:#718096;margin-top:2px">ICH Q8(R2) · USP &lt;1151&gt; · Handbook of Pharmaceutical Excipients 8th Ed.</div>
</div>
""", unsafe_allow_html=True)

# ─── SIDEBAR: HISTORY ────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📋 처방 이력")
    if not st.session_state.history:
        st.info("처방을 생성하면\n이력이 저장됩니다.")
    else:
        for i, h in enumerate(reversed(st.session_state.history)):
            idx = len(st.session_state.history) - i
            with st.expander(f"#{idx} {h['api_name']} {h['api_dose']}mg"):
                st.write(f"**제형:** {h['form_type']}")
                st.write(f"**BCS:** Class {h['bcs']}")
                st.write(f"**API 비율:** {h['api_pct']:.1f}%")
                st.write(f"**정제중량:** {h['weight']}mg")
                st.write(f"**생성:** {h['date']}")

        st.divider()
        hist_json = json.dumps(st.session_state.history, ensure_ascii=False, indent=2)
        st.download_button("📥 이력 JSON 다운로드", data=hist_json,
            file_name=f"AutoForm_history_{datetime.now().strftime('%Y%m%d')}.json",
            mime="application/json", use_container_width=True)
        if st.button("🗑️ 이력 초기화", use_container_width=True):
            st.session_state.history = []
            st.rerun()

    st.divider()
    st.markdown("""
    <div style="font-size:10px;color:#718096;line-height:1.6">
    <b>업데이트 방법</b><br>
    GitHub에 app.py 수정 후 push →<br>
    Streamlit Cloud 자동 반영 ✅<br><br>
    <b>참고 기준</b><br>
    ICH Q8(R2)<br>
    USP &lt;1151&gt;<br>
    HoPE 8th Ed. (Rowe et al.)
    </div>
    """, unsafe_allow_html=True)

# ─── STEP 1: API ──────────────────────────────────────────────────────────────
st.markdown('<div class="section-label">① 주성분 (Active Pharmaceutical Ingredient) 정보  <span style="font-size:10px;font-weight:400;color:#90CDF4">* 필수 입력</span></div>', unsafe_allow_html=True)
with st.container():
    c1, c2 = st.columns(2)
    api_name = c1.text_input("주성분명 (API Name) *", placeholder="예: Metformin HCl, Atorvastatin Calcium")
    api_dose = c2.number_input("목표 함량 (mg/정) *", min_value=0.001, max_value=5000.0, value=None, step=1.0, format="%.3f", placeholder="예: 500")

    c3, c4 = st.columns(2)
    bcs_opts = {"":"— 선택 —","I":"BCS I · 고용해도 / 고투과성 (예: Metformin)","II":"BCS II · 저용해도 / 고투과성 (예: Atorvastatin)","III":"BCS III · 고용해도 / 저투과성 (예: Atenolol)","IV":"BCS IV · 저용해도 / 저투과성 (예: HCT)","unknown":"불명확 / 미분류"}
    bcs = c3.selectbox("BCS 분류 *", list(bcs_opts.keys()), format_func=lambda x: bcs_opts[x])

    with c4:
        st.write("API 물리화학적 특성")
        ca, cb, cc = st.columns(3)
        hygro = ca.checkbox("흡습성")
        light = cb.checkbox("광감수성")
        heat  = cc.checkbox("열감수성")
        cd, ce = st.columns(2)
        ultra_low = cd.checkbox("극저함량 (<1mg)")
        poor_flow = ce.checkbox("유동성 불량")

# ─── STEP 2: TABLET SPEC ──────────────────────────────────────────────────────
st.markdown('<div class="section-label" style="margin-top:6px">② 정제 사양 (Tablet Specification)</div>', unsafe_allow_html=True)
c5, c6, c7 = st.columns(3)
weight_opts = [None,100,150,200,250,300,350,400,500,600,700,800,'직접 입력']
tw_sel = c5.selectbox("목표 정제 중량 (mg) *", weight_opts, format_func=lambda x: "— 선택 —" if x is None else f"{x} mg" if isinstance(x, int) else x)
tablet_weight = c5.number_input("직접 입력 (mg)", 50, 2000, 300) if tw_sel == '직접 입력' else tw_sel
release = c6.selectbox("방출 특성", ["IR (즉방형)","ODT (구강붕해정)","SR (서방형)","EC (장용성)"])
coating = c7.selectbox("코팅 여부", ["비코팅","필름코팅","장용코팅","당의코팅"])

# ─── STEP 3: METHOD ───────────────────────────────────────────────────────────
st.markdown('<div class="section-label" style="margin-top:6px">③ 제형 방법 선택 (Manufacturing Process)</div>', unsafe_allow_html=True)
form_labels = {'DC':'⬛ 직접압축법 (DC)\nDirect Compression\n분말 직타 · 최단 공정','WG':'💧 습식과립법 (WG)\nWet Granulation\n유동성 불량 · 고함량','DG':'🔩 건식과립법 (DG)\nDry Granulation\n수분·열 민감 API','CAP':'💊 경질캡슐 충전\nHard Capsule Fill\n정제 부적합 API'}
form_type = st.radio("제형 방법 *", list(form_labels.keys()), format_func=lambda x: form_labels[x].split('\n')[0], horizontal=True)
st.caption(form_labels[form_type].split('\n',1)[1].replace('\n',' · '))

st.markdown("<br>", unsafe_allow_html=True)

# ─── GENERATE ─────────────────────────────────────────────────────────────────
if st.button("▶  처방 자동 설계 실행 (Generate Formulation)", type="primary", use_container_width=True):
    errs = []
    if not api_name.strip(): errs.append("주성분명을 입력하세요.")
    if not api_dose: errs.append("목표 함량(mg)을 입력하세요.")
    if not bcs: errs.append("BCS 분류를 선택하세요.")
    if not tablet_weight: errs.append("정제 중량을 선택하세요.")
    if api_dose and tablet_weight and isinstance(tablet_weight, (int,float)) and api_dose >= tablet_weight:
        errs.append("API 함량이 정제 중량 이상입니다.")
    for e in errs: st.error(e)

    if not errs:
        params = {'api_name':api_name.strip(),'api_dose':api_dose,'weight':tablet_weight,'bcs':bcs,'release':release,'coating':coating,'hygro':hygro,'light':light,'heat':heat,'ultra_low':ultra_low,'poor_flow':poor_flow,'form_type':form_type}
        api_pct = api_dose / tablet_weight * 100
        eng = {'DC':engine_dc,'WG':engine_wg,'DG':engine_dg,'CAP':engine_cap}
        result = eng[form_type](params, api_pct)

        st.session_state.result_data = {'params':params,'result':result,'api_pct':api_pct}
        st.session_state.history.append({'api_name':api_name.strip(),'api_dose':api_dose,'weight':tablet_weight,'bcs':bcs,'form_type':form_type,'api_pct':round(api_pct,1),'date':datetime.now().strftime('%Y-%m-%d %H:%M')})
        st.rerun()

# ─── RESULTS ─────────────────────────────────────────────────────────────────
if st.session_state.result_data:
    d = st.session_state.result_data
    params, result, api_pct = d['params'], d['result'], d['api_pct']
    comp = result['comp']

    st.divider()
    type_map = {'DC':'직접압축 (DC)','WG':'습식과립 (WG)','DG':'건식과립 (DG)','CAP':'캡슐 충전'}

    # Summary
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("주성분", params['api_name'])
    m2.metric("함량 / 정제중량", f"{params['api_dose']}mg / {params['weight']}mg")
    m3.metric("API 비율 / BCS", f"{api_pct:.1f}% / Class {params['bcs']}")
    m4.metric("제형 방법", type_map[params['form_type']])

    # Process flow
    st.markdown('<div class="stitle">제조 공정 흐름도 (Manufacturing Process Flow)</div>', unsafe_allow_html=True)
    steps = result['steps']
    flow = ' <span class="parrow">→</span> '.join([f'<span class="pstep">{s}</span>' for s in steps])
    st.markdown(f'<div class="pstep-wrap">{flow}</div>', unsafe_allow_html=True)

    # Table + Chart
    st.markdown('<div class="stitle">처방 구성표 (Formulation Composition)</div>', unsafe_allow_html=True)
    col_t, col_c = st.columns([3,1])

    with col_t:
        rows = [{'공정단계':c['step'],'성분명':c['name'],'기능':c['fn'],'규격':c['grade'],'mg/정':f"{c['mg']:.2f}" if c['step']!='캡슐 쉘' else f"{c['mg']:.0f} (별도)",'비율(%)':f"{c['pct']:.2f}%" if c['step']!='캡슐 쉘' else '—'} for c in comp]
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True, height=min(38*len(comp)+60, 420))
        tot_mg = sum(c['mg'] for c in comp if c['step']!='캡슐 쉘')
        tot_pct = sum(c['pct'] for c in comp if c['step']!='캡슐 쉘')
        st.markdown(f"<div style='text-align:right;font-weight:700;font-size:13px;color:#1E2A3A'>합계: {tot_mg:.1f} mg &nbsp;|&nbsp; {tot_pct:.1f}%</div>", unsafe_allow_html=True)

    with col_c:
        pie_c = [c for c in comp if c['step']!='캡슐 쉘']
        labels = [c['name'].split(' (')[0][:22] for c in pie_c]
        values = [c['pct'] for c in pie_c]
        colors = ['#1E2A3A','#3182CE','#63B3ED','#90CDF4','#276749','#68D391','#9B2335','#FC8181','#B7791F','#F6AD55','#4A5568','#A0AEC0']
        fig = go.Figure(go.Pie(labels=labels, values=values, hole=0.4,
            marker=dict(colors=colors[:len(values)], line=dict(color='white',width=2))))
        fig.update_layout(showlegend=True,legend=dict(font=dict(size=9)),margin=dict(t=5,b=5,l=5,r=5),height=300,paper_bgcolor='white')
        st.plotly_chart(fig, use_container_width=True)

    # Warnings
    for w in result['warn']: st.warning(w)

    # Coating / Release notes
    coat_info = {'필름코팅':'Opadry II / Aquacoat 3–5% w/w. HPMC/PVA + 가소제 + TiO₂. 정제 중량 3–5% 증가.','장용코팅':'Eudragit L30D-55 / HPMCP HP-55 10–15%. pH≥5.5 용해. 중량 5–8% 증가.','당의코팅':'Sucrose+Talc+Gum 다층 코팅. 중량 50–100% 증가 가능. 현재 필름코팅으로 대체 추세.'}
    if params['coating'] != '비코팅': st.info(f"🎨 **{params['coating']}:** {coat_info.get(params['coating'],'')}")
    if 'SR' in params['release']: st.info("⏱ **SR:** HPMC K4M/K15M 등 매트릭스 제어제 추가 필요. 별도 SR 처방 설계 권고.")
    if 'ODT' in params['release']: st.info("💧 **ODT:** Mannitol/Erythritol 교체 + PVPP 10–15% 증량 검토. 붕해 ≤60초 목표.")

    # Rationale
    st.markdown('<div class="stitle">처방 설계 근거 (Formulation Rationale)</div>', unsafe_allow_html=True)
    for r in result['rat']:
        st.markdown(f'<div class="rat-box"><strong>📌 {r["t"]}</strong><br>{r["b"]}</div>', unsafe_allow_html=True)

    # Suitability
    st.markdown('<div class="stitle">공정 적합성 평가 (Process Suitability Assessment)</div>', unsafe_allow_html=True)
    sc1, sc2 = st.columns(2)
    for i, s in enumerate(result['suit']):
        col = sc1 if i%2==0 else sc2
        msg = f"**{s['l']}** — {s['d']}"
        if s['s']=='pass': col.success(f"✓ PASS  |  {msg}")
        elif s['s']=='warn': col.warning(f"⚠ CAUTION  |  {msg}")
        else: col.error(f"✗ FAIL  |  {msg}")

    # Compatibility
    st.markdown('<div class="stitle">주요 호환성 고려사항 (Compatibility Considerations)</div>', unsafe_allow_html=True)
    st.dataframe(pd.DataFrame(result['compat']), use_container_width=True, hide_index=True)

    # Disclaimer
    st.info("⚠️ **주의:** 본 처방은 ICH Q8(R2) 기반 **초기 설계 제안(Draft)**입니다. 실제 개발 시 API-부형제 호환성 시험, DoE 최적화, Scale-up 검증, 안정성 시험이 필수입니다.")

    # Downloads
    st.markdown('<div class="stitle">결과 저장 (Export)</div>', unsafe_allow_html=True)
    dl1, dl2, dl3 = st.columns(3)

    excel_buf = export_excel(params, result, api_pct)
    if excel_buf:
        fname = f"AutoForm_{params['api_name'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        dl1.download_button("📥 Excel 다운로드", data=excel_buf, file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    result_json = json.dumps({'params':params,'comp':result['comp'],'api_pct':api_pct,'date':datetime.now().strftime('%Y-%m-%d %H:%M')}, ensure_ascii=False, indent=2)
    dl2.download_button("📋 JSON 다운로드", data=result_json,
        file_name=f"AutoForm_{params['api_name'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.json",
        mime="application/json", use_container_width=True)

    if dl3.button("↩ 처음부터 다시", use_container_width=True):
        st.session_state.result_data = None
        st.rerun()
